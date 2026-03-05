from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def configure_stdio() -> None:
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(errors="replace")


def parse_delimiter(value: str) -> str:
    if len(value) != 1:
        raise argparse.ArgumentTypeError(
            "Разделитель должен быть ровно одним символом."
        )
    return value


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Конвертация Excel-файлов в CSV.")
    parser.add_argument(
        "paths",
        nargs="*",
        default=["."],
        help="Excel-файлы или директории с Excel-файлами.",
    )
    parser.add_argument(
        "-r",
        "--recursive",
        action="store_true",
        help="Рекурсивный поиск файлов внутри директорий.",
    )
    parser.add_argument(
        "-o",
        "--output-dir",
        type=Path,
        default=None,
        help="Директория для сохранения CSV-файлов.",
    )
    parser.add_argument(
        "-a",
        "--all-sheets",
        action="store_true",
        help="Конвертировать все листы книги. Иначе только первый лист.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Имя конкретного листа для конвертации.",
    )
    parser.add_argument(
        "--encoding",
        default="utf-8-sig",
        help="Кодировка выходного CSV (по умолчанию: utf-8-sig).",
    )
    parser.add_argument(
        "-d",
        "--delimiter",
        type=parse_delimiter,
        default=",",
        help="Разделитель в CSV (по умолчанию: ',').",
    )
    return parser.parse_args()


def sanitize_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_")
    return cleaned or "sheet"


def collect_excel_files(paths: list[str], recursive: bool) -> list[Path]:
    files: set[Path] = set()
    for raw_path in paths:
        path = Path(raw_path).expanduser().resolve()
        if not path.exists():
            print(f"Пропуск: путь не найден: {path}")
            continue

        if path.is_file():
            if path.suffix.lower() in SUPPORTED_EXTENSIONS:
                files.add(path)
            else:
                print(f"Пропуск: неподдерживаемое расширение: {path}")
            continue

        if path.is_dir():
            pattern = "**/*" if recursive else "*"
            for candidate in path.glob(pattern):
                if (
                    candidate.is_file()
                    and candidate.suffix.lower() in SUPPORTED_EXTENSIONS
                ):
                    files.add(candidate.resolve())
            continue

        print(f"Пропуск: не файл и не директория: {path}")

    return sorted(files)


def build_output_path(
    source_file: Path,
    sheet_name: str,
    output_dir: Path | None,
    append_sheet_suffix: bool,
) -> Path:
    base_dir = output_dir if output_dir is not None else source_file.parent
    stem = source_file.stem
    if append_sheet_suffix:
        stem = f"{stem}__{sanitize_sheet_name(sheet_name)}"
    return base_dir / f"{stem}.csv"


def write_sheet_to_csv(
    worksheet, target_file: Path, encoding: str, delimiter: str
) -> None:
    target_file.parent.mkdir(parents=True, exist_ok=True)
    with target_file.open("w", newline="", encoding=encoding) as csv_file:
        writer = csv.writer(csv_file, delimiter=delimiter)
        for row in worksheet.iter_rows(values_only=True):
            writer.writerow(["" if value is None else value for value in row])


def convert_excel_file(
    excel_file: Path,
    output_dir: Path | None,
    all_sheets: bool,
    requested_sheet: str | None,
    encoding: str,
    delimiter: str,
) -> int:
    workbook = load_workbook(excel_file, read_only=True, data_only=True)

    if requested_sheet is not None:
        if requested_sheet not in workbook.sheetnames:
            print(f"Ошибка: лист '{requested_sheet}' не найден в {excel_file}")
            workbook.close()
            return 0
        sheet_names = [requested_sheet]
    elif all_sheets:
        sheet_names = workbook.sheetnames
    else:
        sheet_names = [workbook.sheetnames[0]]

    many_outputs = len(sheet_names) > 1
    created = 0

    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        output_path = build_output_path(
            source_file=excel_file,
            sheet_name=sheet_name,
            output_dir=output_dir,
            append_sheet_suffix=many_outputs,
        )
        write_sheet_to_csv(worksheet, output_path, encoding, delimiter)
        print(f"Готово: {excel_file.name} [{sheet_name}] -> {output_path}")
        created += 1

    workbook.close()
    return created


def main() -> int:
    configure_stdio()
    args = parse_args()
    output_dir = args.output_dir.resolve() if args.output_dir else None
    excel_files = collect_excel_files(args.paths, recursive=args.recursive)

    if not excel_files:
        print("Excel-файлы не найдены.")
        return 1

    created_total = 0
    for excel_file in excel_files:
        created_total += convert_excel_file(
            excel_file=excel_file,
            output_dir=output_dir,
            all_sheets=args.all_sheets,
            requested_sheet=args.sheet,
            encoding=args.encoding,
            delimiter=args.delimiter,
        )

    print(f"Всего создано CSV-файлов: {created_total}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
